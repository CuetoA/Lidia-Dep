# -*- coding: utf-8 -*-
"""
Created on Wed Feb 17 13:05:21 2021

@author: Andrés Cueto Estrada
Nombre: Escritura_archivos
Descripción: Este script contendrá funciones de escritura de archivos
Entradas: 
    - Diccionarios de datos
    - Formato de los archivos
    - Objeto tipo archivo
Salidas: - Archivos generados
Comentarios: 
"""
from Biblioteca_objetos import informacion
from Biblioteca_objetos import trabajador
from openpyxl import Workbook
from Lectura_archivos import bandera_columnas_a_leer
from Formatos_archivos import combinar_renglon
from Formatos_archivos import gen_encabezado

def escribiendo_excel(archivo, diccionario):
    # En este caso, diccionario es diccionario_le
    wb = Workbook()
    trabajador_auxiliar = trabajador('','','')
    # Banderas de columnas a escribir
    banderas_columnas = bandera_columnas_a_leer(archivo)    
    
    # Vemos que tipo de diccionario llega
    lista_valores = [ *diccionario.values() ]
    es_diccionario = type(lista_valores[0]) == dict
    es_list = type(lista_valores[0]) == list
    es_string = type(lista_valores[0]) is str
    es_int = type(lista_valores[0]) is int
    es_float = type(lista_valores[0]) is float
    # Creo que esto corrigió el error: NO LO HIZO, ELIMINAR
    es_type = type(lista_valores[0]) is type
    # Creo que esto si lo arregla:
    es_trabajador = type(lista_valores[0]) is type(trabajador_auxiliar)
    
    
    
    if es_diccionario or es_list:
        # Escribir el número de hojas que el número de diccionarios
        contador = 1
        for elemento in diccionario:
            # Dándole el título a la hoja
            if contador == 1:
                print('creando primera sheet llamada {}'.format(elemento))
                ws = wb.active
                ws.title = elemento
            else:
                print('creando siguiente sheet llamada {}'.format(elemento))
                ws = wb.create_sheet(title= elemento)
            contador += 1
            # Escribiendo cada una de las hojas
            ws = escribiendo_hoja(ws, archivo, diccionario[elemento], banderas_columnas)
        wb.save(filename = archivo.ruta)
        return True
        
        
    elif es_string or es_int or es_float or es_type or es_trabajador:
        # Escribir una sola hoja
        ws = wb.active
        #ws = wb.create_sheet(title= archivo.nombre)
        ws.title = archivo.nombre
        ws = escribiendo_hoja(ws, archivo, diccionario, banderas_columnas)
        wb.save(filename = archivo.ruta)
        return True
    else:
        print('Valor {} tipo {} no esperado durante lectura en Escritura_archivos'.format(lista_valores[0],type(lista_valores[0])))
        

    return False

def escribiendo_hoja(ws, archivo, diccionario, bandera_columnas):
    archivo.renglones.actual = 1
    ws = gen_encabezado(ws, archivo, bandera_columnas) 
    ws = combinar_renglon(ws, archivo)

    
    for valor in diccionario.values():
        archivo.siguiente_renglón()
        ws = escribiendo_renglon(ws, archivo, bandera_columnas, valor)
        ws = combinar_renglon(ws, archivo)
    return ws
        
        
def escribiendo_renglon(ws, archivo, banderas_columnas, trabajador):
    #print('columna nombre: ',archivo.columnas.nombre)
    #print('columna categoria: ',archivo.columnas.categoria)
    # seleccionar que columnas se leerán
    renglon = str(archivo.renglones.actual)
    # Condiciones de lectura
    if banderas_columnas[0]: 
        ws[archivo.columnas.numero_empleado + renglon] = str(trabajador.numero_empleado)
        #print('Nombre: \t Col {} \t Numemp{}'.format(archivo.columnas.numero_empleado,trabajador.numero_empleado))
    if banderas_columnas[1]:
        ws[archivo.columnas.nombre + renglon] = str(trabajador.nombre_completo)
        #print('Nombre: \t Col {} \t Nombre{}'.format(archivo.columnas.nombre,trabajador.nombre_completo))
    if banderas_columnas[2]:
        ws[archivo.columnas.categoria + renglon] = str(trabajador.categoria)
        #print('Catego: \t Col {} \t Catego{}'.format(archivo.columnas.categoria,trabajador.categoria))
        
    if banderas_columnas[3]:
        ws[archivo.columnas.hora_entrada + renglon] = str(trabajador.hora_entrada)
        #print('Nombre: \t Col {} \t Entra{}'.format(archivo.columnas.hora_entrada,trabajador.hora_entrada))
    if banderas_columnas[4]:
        ws[archivo.columnas.hora_salida + renglon] = str(trabajador.hora_salida)
        #print('Nombre: \t Col {} \t Salid{}'.format(archivo.columnas.hora_salida,trabajador.hora_salida))
    if banderas_columnas[5]:
        ws[archivo.columnas.retardo + renglon] = str(trabajador.retardo)
        #print('Nombre: \t Col {} \t Retard{}'.format(archivo.columnas.retardo,trabajador.retardo))
    if banderas_columnas[6]:
        ws[archivo.columnas.tiempo_ext + renglon] = str(trabajador.tiempo_extra)
        #print('Nombre: \t Col {} \t tiempex{}'.format(archivo.columnas.tiempo_ext,trabajador.tiempo_extra))
    if banderas_columnas[7]:
        ws[archivo.columnas.jornada + renglon] = str(trabajador.jornada_laboral)
        #print('Nombre: \t Col {} \t jornada{}'.format(archivo.columnas.jornada,trabajador.jornada_laboral))
    if banderas_columnas[8]:
        ws[archivo.columnas.observaciones + renglon] = str(trabajador.observaciones)
        #print('Nombre: \t Col {} \t obser{}'.format(archivo.columnas.observaciones,trabajador.observaciones))
    if banderas_columnas[9]:
        ws[archivo.columnas.superior + renglon] = str(trabajador.superior)
        #print('Nombre: \t Col {} \t superi{}'.format(archivo.columnas.superior,trabajador.superior))
        #print()
    return ws
    

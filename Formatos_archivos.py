# -*- coding: utf-8 -*-
"""
Created on Wed Feb 17 13:19:12 2021

@author: Andrés Cueto Estrada
Nombre: Formatos_archivos, cambiara el nombre a 
Descripción: Este script contendrá formatos de escritura de archivos
Entradas: 
    - Formato de los archivos
    - Objeto tipo archivo
Salidas: - Archivos generados
Comentarios: 
"""
from Biblioteca_objetos import informacion
from openpyxl import Workbook



def gen_encabezado(ws, archivo, banderas_columnas):
    # ELIMINAR las siguientes dos líneas inmediatas antes de ejecutar
    info = informacion
    archivo =  info.bd
    renglon = '1'
        
    # Condiciones de lectura
    if banderas_columnas[0]: 
        ws[archivo.columnas.numero_empleado + renglon] = archivo.columnas.numero_empleado_tit
    if banderas_columnas[1]:
        ws[archivo.columnas.nombre + renglon] = archivo.columnas.nombre_tit
    if banderas_columnas[2]:
        ws[archivo.columnas.categoria + renglon] = archivo.columnas.categoria_tit
    if banderas_columnas[3]:
        ws[archivo.columnas.hora_entrada + renglon] = archivo.columnas.hora_entrada_tit
    if banderas_columnas[4]:
        ws[archivo.columnas.hora_salida + renglon] = archivo.columnas.hora_salida_tit
    if banderas_columnas[5]:
        ws[archivo.columnas.retardo + renglon] = archivo.columnas.retardo_tit
    if banderas_columnas[6]:
        ws[archivo.columnas.tiempo_ext + renglon] = archivo.columnas.tiempo_ext_tit
    if banderas_columnas[7]:
        ws[archivo.columnas.jornada + renglon] = archivo.columnas.jornada_tit
    if banderas_columnas[8]:
        ws[archivo.columnas.observaciones + renglon] = archivo.columnas.observaciones_tit
    if banderas_columnas[9]:
        ws[archivo.columnas.superior + renglon] = archivo.columnas.superior_tit
    
    # Renglón de anotaciones generales    
    if archivo.nombre == informacion.bd.nombre:
        ws, archivo = celda_general_anotaciones(ws, archivo)
    
    return ws

def celda_general_anotaciones(ws, archivo):
    archivo.renglones.actual = 2
    renglon = str(archivo.renglones.actual)
    # ELIMINAR comentario
    #print('renglon: ',renglon)
    
    ws[archivo.columnas.numero_empleado + renglon] = '###'
    ws[archivo.columnas.categoria + renglon] = '###'
    ws[archivo.columnas.nombre + renglon] = 'Anotaciones Generales'.capitalize()
    
    
    
    return ws, archivo
        
def combinar_renglon(ws, archivo):
    
    # Aquí debemos poner un condicional
    ws = combinar_celda(ws, archivo.columnas.nombre_comb, archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.superior_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.categoria_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.numero_empleado_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.hora_entrada_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.hora_salida_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.retardo_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.tiempo_ext_comb , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.jornada_comb  , archivo.renglones.actual)
    ws = combinar_celda(ws, archivo.columnas.observaciones_comb  , archivo.renglones.actual)
    return ws
    
    
def combinar_celda(ws, lista_columnas, renglon):
    if len(lista_columnas ) > 1:
        celda_inicio = lista_columnas[0] + str(renglon)
        celda_final = lista_columnas[1] + str(renglon)
        combinacion = celda_inicio + ':' + celda_final
        ws.merge_cells(combinacion)
    return ws






# -*- coding: utf-8 -*-
"""
Created on Mon Feb 15 22:40:08 2021

@author: Andrés Cueto Estrada
Nombre: Lectura_archivos
Descripción: 
    Este script tendrá todas las funciones de lectruda, esto con el objetivo de 
    reutilizar ciertas funciones a nivel local del script
Entradas: - Triggers
Salidas: 
    - Listas de empleados, base de datos, etc...
Comentarios: 
"""
import openpyxl 
from Biblioteca_objetos import informacion
from Biblioteca_objetos import trabajador
from openpyxl import Workbook

def lectura_documentos (bandera, archivo):
    # Declaramos los objetos de salida que tendremos
    diccionario = {}
    # Rumbo de acción
    if not bandera: return False, diccionario
        
    # Abrir el Wb 
    wb = openpyxl.load_workbook(archivo.ruta)
    # Obtener las banderas de lectura
    banderas = bandera_columnas_a_leer(archivo)
    # Iterando entre hojas
    hojas = wb.sheetnames
    for hoja in hojas:
        
        # Leer hoja bd
        ws = wb[hoja]
        
        diccionario = lectura_hoja_bd(ws, archivo, diccionario, banderas)
        # Después abseorveremos el diccionario generado en lectura hoja bd
    
    return True, diccionario
    

def lectura_hoja_bd(ws, archivo, diccionario, b_cel):
    # b_cel : bandera de celdas
    # Inicializamos contadores
    espacios_en_blanco = 0 
    sin_matricula_contador = 0
    
    # Leeremos el renglón de anotaciones generales
    esquema_general = lectura_renglon_anotaciones(ws, archivo)
    
    
    
    # Iteramos entre los renglones obteniendo los datos
    while espacios_en_blanco < 10:
        archivo.siguiente_renglón()
        #archivo.generando_celdas_actuales()
        
        diccionario, espacios_en_blanco, sin_matricula_contador = lectura_renglon(ws, archivo, diccionario, b_cel, espacios_en_blanco, sin_matricula_contador, esquema_general)
    
    return diccionario


def lectura_renglon_anotaciones(ws, archivo):
    # Solo se aplicará si es bd, sino, retornará con funciones vacías
    if archivo is not informacion.la1:
        gral_entrada_str = ''
        gral_salida_str = ''
    
        bandera_gral_entrada = False
        bandera_gral_salida = False
        lista_entrada = [bandera_gral_entrada, gral_entrada_str]
        lista_salida = [bandera_gral_salida,gral_salida_str]
        
        esquema_general = {'entrada': lista_entrada , 'salida': lista_salida}
        return esquema_general
    
    
    renglon = str(archivo.renglones.anotacion_general)
    col_entrada = archivo.columnas.hora_entrada
    col_salida = archivo.columnas.hora_salida
    
    print('col_entrada: \t', col_entrada)
    print('col_salida: \t', col_salida)
    print('renglon: \t', renglon)
    
    gral_entrada_str = str(ws[col_entrada + renglon].value)
    gral_salida_str = str(ws[col_salida + renglon].value)
    
    bandera_gral_entrada = gral_entrada_str != ''
    bandera_gral_salida = gral_salida_str != ''
    
    lista_entrada = [bandera_gral_entrada, gral_entrada_str]
    lista_salida = [bandera_gral_salida,gral_salida_str]
    
    esquema_general = {'entrada': lista_entrada , 'salida': lista_salida}
    
    return esquema_general
    


def lectura_renglon(ws, archivo, diccionario_bd, b_cel, espacios_en_blanco, sin_matricula_contador, esquema_general):
    # Instanciando objeto trabajador
    trabajador_n = trabajador('', '', '')
    
    
    # Desempacando banderas y esquemas de trabajo
    lista_entrada = esquema_general['entrada']
    lista_salida = esquema_general['salida']
    
    bandera_gral_entrada = lista_entrada[0]
    gral_entrada_str = lista_entrada[1]
    bandera_gral_salida = lista_salida[0]
    gral_salida_str = lista_salida[1]
    
    #  Obteneiendo los esquemas generales del empleado
    trabajador_n.aplicar_esuqema_general_entrada = bandera_gral_entrada
    trabajador_n.aplicar_esuqema_general_salida = bandera_gral_salida
    trabajador_n.esquema_gral_entrada = gral_entrada_str
    trabajador_n.esquema_gral_salida = gral_salida_str
    
    
    # Declaración banderas de vacio
    celda_numero_empleado_vacia = True
    celda_nombre_vacia = True
    celda_categoria_vacia = True
    celda_hora_entrada_vacia = True
    celda_hora_salida_vacia = True
    celda_retardo_vacia = True
    celda_tiempo_extr_vacia = True
    celda_jornada_vacia = True
    celda_observaciones_vacia = True
    celda_superior_vacia = True
    
    
    # Obteniendo los datos del empleado según sus Condiciones de lectura
    if b_cel[0]: 
        celda_numero_empleado = ws[archivo.celdas.numero_empleado]
        celda_numero_empleado_vacia = celda_numero_empleado.value is None
        if not celda_numero_empleado_vacia: trabajador_n.numero_empleado = str(celda_numero_empleado.value)
    if b_cel[1]:
        celda_nombre = ws[archivo.celdas.nombre]
        celda_nombre_vacia = celda_nombre.value is None
        if not celda_nombre_vacia: trabajador_n.nombre_completo = str(celda_nombre.value)
    if b_cel[2]:
        celda_categoria = ws[archivo.celdas.categoria]
        celda_categoria_vacia = celda_categoria.value is None
        if not celda_categoria_vacia: trabajador_n.categoria = str(celda_categoria.value)
    if b_cel[3]:
        celda_hora_entrada = ws[archivo.celdas.hora_entrada]
        celda_hora_entrada_vacia = celda_hora_entrada.value is None
        if not celda_hora_entrada_vacia: trabajador_n.hora_entrada = str(celda_hora_entrada.value)
    if b_cel[4]:
        celda_hora_salida = ws[archivo.celdas.hora_salida]
        celda_hora_salida_vacia = celda_hora_salida.value is None
        if not celda_hora_salida_vacia: trabajador_n.hora_salida = str(celda_hora_salida.value)
    if b_cel[5]:
        celda_retardo = ws[archivo.celdas.retardo]
        celda_retardo_vacia = celda_retardo.value is None
        if not celda_retardo_vacia: trabajador_n.retardo = str(celda_retardo.value)
    if b_cel[6]:
        celda_tiempo_extr = ws[archivo.celdas.tiempo_extr] 
        celda_tiempo_extr_vacia = celda_tiempo_extr.value is None
        if not celda_tiempo_extr_vacia: trabajador_n.tiempo_extr = str(celda_tiempo_extr.value)
    if b_cel[7]:
        celda_jornada = ws[archivo.celdas.jornada] 
        celda_jornada_vacia = celda_jornada.value is None
        if not celda_jornada_vacia: trabajador_n.jornada = str(celda_jornada.value)
    if b_cel[8]:
        celda_observaciones = ws[archivo.celdas.observaciones] 
        celda_observaciones_vacia = celda_observaciones.value is None
        if not celda_observaciones_vacia: trabajador_n.observaciones = str(celda_observaciones.value)
    if b_cel[9]:
        celda_superior = ws[archivo.celdas.superior] 
        celda_superior_vacia = celda_superior.value is None
        if not celda_superior_vacia: trabajador_n.superior = str(celda_superior.value)
    
    # si todas las celdas están vacías se aumenta el contador de espacios vacíos
    bandera1 = celda_numero_empleado_vacia and celda_nombre_vacia and celda_categoria_vacia
    bandera2 = celda_hora_entrada_vacia and celda_hora_salida_vacia and celda_retardo_vacia
    bandera3 = celda_tiempo_extr_vacia and celda_jornada_vacia and celda_observaciones_vacia
    todas_celdas_vacias = bandera1 and bandera2 and bandera3 and celda_superior_vacia
    if todas_celdas_vacias:
        espacios_en_blanco += 1
        return diccionario_bd, espacios_en_blanco, sin_matricula_contador
    else:
        espacios_en_blanco = 0
    
    # Guardar el dato en el diccionario
    if celda_numero_empleado_vacia:
        sin_matricula_contador += 1
        llave = 'Sin matricula ' + str(sin_matricula_contador)
    else:
        llave = trabajador_n.numero_empleado
    valor = trabajador_n
    diccionario_bd[llave] = valor
    
    trabajador_n.clear
    trabajador.clear
    
    return diccionario_bd, espacios_en_blanco, sin_matricula_contador

    
  
    
def bandera_columnas_a_leer(archivo):
    # Esta función deberá detectar que columnas si se leerán, segú estén vacías o no
    # Declarando las banderas que habrá
    b_numero_empleado = False
    b_nombre = False
    b_categoria = True
    b_hora_entrada = False
    b_hora_salida = False
    b_retardo = False
    b_tiempo_ext = False
    b_jornada = False
    b_observaciones = False
    b_superior = False
    
    # Obteniendo las banderas
    if len(archivo.columnas.numero_empleado) > 0: b_numero_empleado = True
    if len(archivo.columnas.nombre) > 0: b_nombre = True
    if len(archivo.columnas.categoria) > 0: b_categoria = True
    if len(archivo.columnas.hora_entrada) > 0: b_hora_entrada = True
    if len(archivo.columnas.hora_salida) > 0: b_hora_salida = True
    if len(archivo.columnas.retardo) > 0: b_retardo = True
    if len(archivo.columnas.tiempo_ext) > 0: b_tiempo_ext = True
    if len(archivo.columnas.jornada) > 0: b_jornada = True
    if len(archivo.columnas.observaciones) > 0: b_observaciones = True
    if len(archivo.columnas.superior) > 0: b_superior = True
    
    banderas = []
    banderas.append(b_numero_empleado)
    banderas.append(b_nombre)
    banderas.append(b_categoria)
    banderas.append(b_hora_entrada)
    banderas.append(b_hora_salida)
    banderas.append(b_retardo)
    banderas.append(b_tiempo_ext)
    banderas.append(b_jornada)
    banderas.append(b_observaciones)
    banderas.append(b_superior)
    
    return banderas
    


    
    
    
    
    
    
    

